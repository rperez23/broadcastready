#! /usr/bin/python3

import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import os
import pandas as pd
import re
import subprocess
import sys
import warnings

warnings.simplefilter(action='ignore', category=UserWarning)

#Conditions for broadcast ready
#1: hn assigned
#2: hn.scc file on s3
#3: mxf
#4: starts at hour 1

xlrow = 2
wslist = ['Full Ingest Summary', 'Captions Summary']
validformats = ['MXF', 'QuickTime']
housenumbers = {}


def usage():
	print('')
	print('Usage:')
	print('')
	print('   broadcastready.py <xlf name>')
	print('')

def getxlf():

	xlf = sys.argv[1]

	if os.path.isfile(xlf):
		return xlf

	print('')
	print(xlf,'Does not exist, exiting')
	usage()
	sys.exit(1)

def gethousenumbers():

	hnlist = {}

	print('')
	print('ENTER YOUR HOUSE NUMBERS:')
	print('')

	while(1):

		hn = input()

		if hn == '':
			break

		m  = re.match('^BUZ_[A-Z0-9]+$',hn)

		if m:
			hnlist[hn] = []

	if len(hnlist) == 0:
		print('')
		print('No Valid House Numbers, Exiting')
		print('')
		sys.exit(1)


	return hnlist


def getindexes(hn,db,keyval):

	indexlist = []

	for key in db[keyval]:
		if db[keyval][key] == hn:
			indexlist.append(key)

	return indexlist


def printviddata(hn,videodb,vidindexnums,capindexnums,sheetout,xlrow):
	

	episode = ''.ljust(40)
	tc      = '--:--:--;--'
	capf    = ''.ljust(50)
	sccf    = ''.ljust(20)

	if len(capindexnums) > 0:
		sccf = hn + '.scc'
		sccf = sccf.ljust(15)
		
	if len(vidindexnums) == 0:
		print(hn,':',tc,':',episode,':',sccf,':',capf)
		return

	for i in vidindexnums:

		episode      = videodb['Supplier.OriginalName'][i].ljust(40)
		assetid      = videodb['Resource.Name'][i]
		masterformat = videodb['Format.MasterStandard'][i]
		#print(assetid,episode)

		cmd = 'getassetidinfo.py ' + assetid

		status = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True, shell=True).stdout.strip("\n")
		status = status.replace('{','')
		status = status.replace('}','')
		status = status.replace('"','')
		#status = status.replace('}','')

		parts  = status.split(',')
		#print(parts)
		"""
		tc     = parts[3].replace('Format.TimeStart:','')

		capf   = parts[4].replace('TWK.AncillaryName:','').ljust(50)
		print(hn,':',tc,':',episode,':',sccf,':',capf)
		"""

		mtc  = re.search("Format.TimeStart: (.+)$",parts[4])
		mcap = re.search("TWK.AncillaryName: (.+)$",parts[6])

		if mtc:
			tc = mtc.group(1)
		if mcap:
			capf = mcap.group(1).ljust(50)


		greenfill = PatternFill(start_color='90EE90',end_color='90EE90',fill_type='solid')
		redfill   = PatternFill(start_color='FFCCCC',end_color='FFCCCC',fill_type='solid')

		#print(hn,':',tc,':',episode,':',sccf,':',capf)
		print(hn,':',tc,':',masterformat,':',sccf,':',capf)
		sheetout.cell(row=xlrow,column=1).value = hn
		sheetout.cell(row=xlrow,column=2).value = tc 

		if tc == '01:00:00;00' or '01:00:00:00':
			sheetout.cell(row=xlrow,column=2).fill = greenfill
		else:
			sheetout.cell(row=xlrow,column=2).fill = redfill


		sheetout.cell(row=xlrow,column=3).value = masterformat

		if masterformat in validformats:
			sheetout.cell(row=xlrow,column=3).fill = greenfill
		else:
			sheetout.cell(row=xlrow,column=3).fill = redfill

		ms3cap = re.match('^BUZ_',sccf)
		if ms3cap:
			txt = 'YES'
			sheetout.cell(row=xlrow,column=4).value = txt
			sheetout.cell(row=xlrow,column=4).fill = greenfill


		else:
			txt = 'NO'
			sheetout.cell(row=xlrow,column=4).value = txt
			sheetout.cell(row=xlrow,column=4).fill = redfill

		







		#txt = hn + ',' + tc + ',' + masterformat + ',' + sccf + ',' + capf
		#print(txt)





if len(sys.argv) <= 1:
	usage()
	sys.exit(1)


xlf = getxlf()
housenumbers = gethousenumbers()

#open the workbook for reading
try:
	wb = openpyxl.load_workbook(filename=xlf,read_only=True)
except:
	print("  Cannot open",xlf)
	sys.exit(1)

#validate the worksheets are in the xlf
#validate that sheet exists
for tab in wslist:
	if not (tab in wb.sheetnames):
		print(" ",tab,"not in",xlf)
		print('')
		wb.close()
		sys.exit(1)
wb.close()

try:
	wbout = openpyxl.Workbook()
except:
	print('   ~~~Cannot create status xl workbook~~~')
	sys.exit(1)
try:
	sheetout = wbout.active
except:
	print('   ~~~Cannot create status xl sheet~~~')
	sys.exit(1)

videodb   = {}
captiondb = {}

df = pd.read_excel(xlf,sheet_name=wslist[0])
videodb = df.to_dict()

df = pd.read_excel(xlf,sheet_name=wslist[1])
captiondb = df.to_dict()

sheetout.cell(row=1,column=1).value = 'HOUSE NUMBER'
sheetout.cell(row=1,column=2).value = 'START TIME'
sheetout.cell(row=1,column=3).value = 'FORMAT'
sheetout.cell(row=1,column=4).value = 'S3 HN CAPTION'
#sheetout.cell(row=1,column=5).value = 'ASSET CAPTION'


for hn in housenumbers:

	vidindexnums = getindexes(hn,videodb,'Fremantle.HouseNumber')
	#print(vidindexnums)

	capindexnums = getindexes(hn,captiondb,'Supplier.Source')
	#print(capindexnums)

	printviddata(hn,videodb,vidindexnums,capindexnums,sheetout,xlrow)
	xlrow += 1
print('')


sheetout.column_dimensions['A'].width = 15
sheetout.column_dimensions['B'].width = 15
sheetout.column_dimensions['C'].width = 15
sheetout.column_dimensions['D'].width = 15


wbout.save('Broadcast-Ready-Status.xlsx')
wbout.close()	

#Conditions for broadcast ready
#1: hn assigned
#2: hn.scc file on s3
#3: mxf
#4: starts at hour 1




